"""Build AlphaAgent_Financial_Model.xlsx — banker-ready, all-live-formulas.

Reproduces model.py exactly, with straight-line depreciation of the glasses
(3-yr life) so the P&L is smooth while cash still reflects up-front capex.
Every downstream cell is a formula referencing the Assumptions tab.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/nathanaelguitar/Downloads/AlphaAgent_Financial_Model.xlsx"

NAVY = "0E223B"; TEAL = "14B8A6"; YELLOW = "FFF7D6"; GREY = "EEF1F4"; WHITE = "FFFFFF"
CUR = '$#,##0;[Red]($#,##0)'      # currency, red parens for negatives
CUR2 = '$#,##0.00'
PCT = '0%'
thin = Side(style="thin", color="D0D7DE")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

def style(cell, *, bold=False, size=11, color="242E3A", fill=None, fmt=None,
          align=None, wrap=False, border=False):
    cell.font = Font(name="Calibri", size=size, bold=bold, color=color)
    if fill: cell.fill = PatternFill("solid", fgColor=fill)
    if fmt: cell.number_format = fmt
    if align or wrap: cell.alignment = Alignment(horizontal=align, wrap_text=wrap, vertical="center")
    if border: cell.border = box

wb = openpyxl.Workbook()

# ==================== ASSUMPTIONS ====================
a = wb.active; a.title = "Assumptions"
a.sheet_view.showGridLines = False
a.column_dimensions["A"].width = 30
a.column_dimensions["B"].width = 16
a.column_dimensions["C"].width = 52

a.merge_cells("A1:C1")
style(a["A1"], bold=True, size=16, color=WHITE, fill=NAVY)
a["A1"] = "AlphaAgent Healthcare Solutions — Financial Model (Confidential)"
a.merge_cells("A2:C2")
style(a["A2"], size=10, color="64748B")
a["A2"] = "Bottom-up 3-year build · site-priced B2B SaaS medtech · edit the yellow cells; every tab recalculates"
a.row_dimensions[1].height = 26

def section(row, text):
    a.merge_cells(f"A{row}:C{row}")
    style(a[f"A{row}"], bold=True, size=11, color=WHITE, fill=TEAL)
    a[f"A{row}"] = text

def driver(row, label, value, note, fmt=CUR, calc=False):
    style(a[f"A{row}"], bold=True, border=True)
    a[f"A{row}"] = label
    style(a[f"B{row}"], border=True, fmt=fmt, align="right",
          fill=GREY if calc else YELLOW, bold=calc)
    a[f"B{row}"] = value
    style(a[f"C{row}"], size=9, color="64748B", wrap=True)
    a[f"C{row}"] = note

section(4, "UNIT ECONOMICS (drivers)")
driver(5, "Price / site / month", 10000, "Blended site license ($5K entry service line → $15–20K expanded)")
driver(6, "Cloud COGS / site / month", 1000, "Inference, hosting, support")
driver(7, "Glasses per new site", 10, "≈10 surgeons per site", fmt="0")
driver(8, "Cost per glass (LX1)", 2000, "Hardware unit cost")
driver(9, "Hardware / new site", "=B7*B8", "glasses × cost/glass, capitalized when a site onboards", calc=True)
driver(10, "Useful life (months)", 36, "Straight-line depreciation period for the glasses", fmt="0")
driver(11, "Depreciation / site / month", "=B9/B10", "= hardware ÷ useful life; the only hardware cost the P&L sees", fmt=CUR2, calc=True)

section(13, "FUNDING")
driver(14, "Pre-seed (starting cash)", 500000, "Closes month 1 (H2 2026)")
driver(15, "Seed round", 1500000, "Raised on converted pilots")
driver(16, "Seed month", 14, "Month the seed lands", fmt="0")

section(18, "MARKET")
driver(19, "US addressable sites", 17100, "≈6,100 hospitals + ≈11,000 ASCs (for penetration math)", fmt="#,##0")

# ---- monthly schedule (editable) ----
section(21, "MONTHLY SCHEDULE (editable inputs — cumulative paying sites & opex)")
hdr = 22
for col, txt in [("A", "Month"), ("B", "Cumulative Paying Sites"), ("C", "Monthly Opex")]:
    style(a[f"{col}{hdr}"], bold=True, color=WHITE, fill=NAVY, align="center", border=True)
    a[f"{col}{hdr}"] = txt
sites = [0,0,0,0,0,0,0,0,1,2,3,3, 4,4,5,6,6,7,8,8,9,9,10,10,
         12,13,14,15,16,18,19,20,22,23,24,25]
opex = [29000]*12 + [60000]*6 + [85000]*6 + [110000]*6 + [140000]*6
SCHED0 = hdr + 1                      # first schedule data row = 23
for i in range(36):
    r = SCHED0 + i
    style(a[f"A{r}"], align="center", border=True); a[f"A{r}"] = i + 1
    style(a[f"B{r}"], align="center", border=True, fill=YELLOW); a[f"B{r}"] = sites[i]
    style(a[f"C{r}"], align="right", border=True, fill=YELLOW, fmt=CUR); a[f"C{r}"] = opex[i]

# ---- opex note ----
note_row = SCHED0 + 36 + 1            # 60
section(note_row, "OPEX NOTE — two founders, not a growing team")
a.merge_cells(f"A{note_row+1}:C{note_row+4}")
style(a[f"A{note_row+1}"], size=10, wrap=True, color="242E3A")
a[f"A{note_row+1}"] = (
    "Opex is the two founders + non-headcount costs, NOT expanding staff. "
    "Y1 (~$29K/mo): Nathanael's salary, base cloud, regulatory setup. "
    "Y2 step-up: Manny moves from equity-only to salaried (the ONLY headcount change) as 510(k) work ramps. "
    "Y3 step-up: 510(k) submission costs, medical-device / D&O insurance, cloud at scale, and 1099 "
    "implementation & sales contractors (not FTEs). The company is designed to run on the two founders "
    "indefinitely — the ramp is regulatory, infrastructure, and go-to-market spend, not payroll."
)
a.row_dimensions[note_row+1].height = 70

# handy absolute refs
P   = "Assumptions!$B$5"
CL  = "Assumptions!$B$6"
HW  = "Assumptions!$B$9"
DEP = "Assumptions!$B$11"
PRESEED = "Assumptions!$B$14"
SEED = "Assumptions!$B$15"
SEEDM = "Assumptions!$B$16"
USSITES = "Assumptions!$B$19"

# ==================== MONTHLY MODEL ====================
m = wb.create_sheet("Monthly Model")
m.sheet_view.showGridLines = False
m.merge_cells("A1:N1")
style(m["A1"], bold=True, size=14, color=WHITE, fill=NAVY)
m["A1"] = "Monthly Model — 36-month bottom-up build (all cells live formulas → Assumptions)"
m.row_dimensions[1].height = 22

cols = ["Month", "Year", "Paying Sites", "New Sites", "Revenue", "Cloud COGS",
        "Depreciation", "Total COGS", "Gross Profit", "GM %", "Opex",
        "Net Income", "Capex (cash)", "Cash Balance"]
HR = 3
for j, c in enumerate(cols):
    cell = m.cell(HR, j + 1, c)
    style(cell, bold=True, color=WHITE, fill=NAVY, align="center", border=True, wrap=True)
widths = [7, 6, 12, 10, 12, 12, 12, 12, 12, 8, 12, 12, 12, 14]
for j, w in enumerate(widths):
    m.column_dimensions[get_column_letter(j + 1)].width = w
m.freeze_panes = "A4"

D0 = HR + 1                           # first data row = 4
for i in range(36):
    r = D0 + i
    sched = SCHED0 + i                # Assumptions row for this month
    prev = r - 1
    m.cell(r, 1, i + 1)                                            # Month
    m.cell(r, 2, f"=ROUNDUP(A{r}/12,0)")                          # Year
    m.cell(r, 3, f"=Assumptions!$B${sched}")                      # Paying Sites
    m.cell(r, 4, f"=C{r}" if i == 0 else f"=C{r}-C{prev}")        # New Sites
    m.cell(r, 5, f"=C{r}*{P}")                                    # Revenue
    m.cell(r, 6, f"=C{r}*{CL}")                                   # Cloud COGS
    m.cell(r, 7, f"=C{r}*{DEP}")                                  # Depreciation
    m.cell(r, 8, f"=F{r}+G{r}")                                   # Total COGS
    m.cell(r, 9, f"=E{r}-H{r}")                                   # Gross Profit
    m.cell(r, 10, f"=IF(E{r}=0,0,I{r}/E{r})")                     # GM %
    m.cell(r, 11, f"=Assumptions!$C${sched}")                    # Opex
    m.cell(r, 12, f"=I{r}-K{r}")                                  # Net Income
    m.cell(r, 13, f"=D{r}*{HW}")                                  # Capex
    seed = f"+IF(A{r}={SEEDM},{SEED},0)"
    if i == 0:
        m.cell(r, 14, f"={PRESEED}+L{r}+G{r}-M{r}{seed}")        # Cash (op CF + back dep - capex)
    else:
        m.cell(r, 14, f"=N{prev}+L{r}+G{r}-M{r}{seed}")
    # formatting
    for j, fmt in [(1,"0"),(2,"0"),(3,"0"),(4,"0"),(5,CUR),(6,CUR),(7,CUR),
                   (8,CUR),(9,CUR),(10,PCT),(11,CUR),(12,CUR),(13,CUR),(14,CUR)]:
        cell = m.cell(r, j)
        style(cell, border=True, fmt=fmt, align="center" if j <= 4 else "right",
              fill=WHITE if i % 2 == 0 else GREY)

# ==================== ANNUAL SUMMARY ====================
s = wb.create_sheet("Annual Summary")
s.sheet_view.showGridLines = False
s.merge_cells("A1:D1")
style(s["A1"], bold=True, size=14, color=WHITE, fill=NAVY)
s["A1"] = "Annual Summary — rolls up the Monthly Model (live)"
for col, w in [("A", 26), ("B", 15), ("C", 15), ("D", 15)]:
    s.column_dimensions[col].width = w
for j, yr in enumerate(["", "Year 1", "Year 2", "Year 3"]):
    style(s.cell(3, j + 1, yr), bold=True, color=WHITE, fill=NAVY,
          align="right" if j else "left", border=True)

MM = "'Monthly Model'"
def yrcol(letter, y): return f"SUMIF({MM}!$B$4:$B$39,{y},{MM}!${letter}$4:${letter}$39)"
def eoy(letter, y):   return f"INDEX({MM}!${letter}$4:${letter}$39,MATCH({y}*12,{MM}!$A$4:$A$39,0))"

srows = [
    ("Revenue",       lambda y: f"={yrcol('E',y)}", CUR),
    ("Cloud COGS",    lambda y: f"={yrcol('F',y)}", CUR),
    ("Depreciation",  lambda y: f"={yrcol('G',y)}", CUR),
    ("Total COGS",    lambda y: f"={yrcol('H',y)}", CUR),
    ("Gross Profit",  lambda y: f"={yrcol('I',y)}", CUR),
    ("Gross Margin %",lambda y: f"={yrcol('I',y)}/{yrcol('E',y)}", PCT),
    ("Opex",          lambda y: f"={yrcol('K',y)}", CUR),
    ("Net Income",    lambda y: f"={yrcol('L',y)}", CUR),
    ("Capex (cash)",  lambda y: f"={yrcol('M',y)}", CUR),
    ("Sites (EOY)",   lambda y: f"={eoy('C',y)}", "0"),
    ("Exit ARR",      lambda y: f"={eoy('C',y)}*{P}*12", CUR),
    ("End-of-year Cash", lambda y: f"={eoy('N',y)}", CUR),
]
for i, (label, fn, fmt) in enumerate(srows):
    r = 4 + i
    bold = label in ("Net Income", "Exit ARR", "Gross Profit")
    style(s.cell(r, 1, label), bold=bold, border=True,
          fill=GREY if i % 2 else WHITE)
    for y in (1, 2, 3):
        cell = s.cell(r, 1 + y, fn(y))
        style(cell, border=True, fmt=fmt, align="right", bold=bold,
              fill=GREY if i % 2 else WHITE)

s.merge_cells("A17:D17")
style(s["A17"], size=9, color="64748B", wrap=True)
s["A17"] = ("Glasses are capitalized and depreciated straight-line over 36 months, so Net Income shows "
            "depreciation (non-cash) while Capex shows the full hardware cash outlay — that is why End-of-year "
            "Cash reflects the real $2M of capital deployed even though the P&L smooths out.")
s.row_dimensions[17].height = 44

# ==================== LONG VIEW ====================
lv = wb.create_sheet("Long View")
lv.sheet_view.showGridLines = False
for col, w in [("A", 12), ("B", 10), ("C", 18), ("D", 18), ("E", 18)]:
    lv.column_dimensions[col].width = w
lv.merge_cells("A1:E1")
style(lv["A1"], bold=True, size=13, color=WHITE, fill=NAVY)
lv["A1"] = "Long View — illustrative scenario (SEPARATE from the bottom-up 3-yr build)"
lv.merge_cells("A2:E2")
lv["A2"] = "Post-510(k) departmental expansion. Sites & Avg ACV are editable scenario inputs; ARR and penetration are formulas."
lv["A2"].font = Font(name="Calibri", size=9, italic=True, color="64748B")
for j, h in enumerate(["Year", "Sites", "Avg ACV / site", "ARR", "US Penetration"]):
    style(lv.cell(3, j + 1, h), bold=True, color=WHITE, fill=NAVY,
          align="center" if j else "left", border=True)
lvdata = [(3, 25, 120000), (5, 90, 150000), (7, 260, 165000), (10, 750, 180000)]
for i, (yr, sites_n, acv) in enumerate(lvdata):
    r = 4 + i
    style(lv.cell(r, 1, f"Year {yr}"), bold=True, border=True)
    style(lv.cell(r, 2, sites_n), align="center", border=True, fill=YELLOW)
    style(lv.cell(r, 3, acv), align="right", border=True, fmt=CUR, fill=YELLOW)
    style(lv.cell(r, 4, f"=B{r}*C{r}"), align="right", border=True, fmt=CUR, bold=True, color=TEAL)
    style(lv.cell(r, 5, f"=B{r}/{USSITES}"), align="right", border=True, fmt="0.0%")

wb.save(OUT)
print("saved", OUT)
